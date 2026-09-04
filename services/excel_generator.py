from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


class ExcelGenerator:

    def generate(
            self,
            sheets: list[dict],
            output_path: str,
    ):
        workbook = Workbook()

        # Remove default sheet
        default_sheet = workbook.active
        workbook.remove(default_sheet)

        for sheet_data in sheets:
            sheet_name = sheet_data.get(
                "name",
                "Sheet",
            )

            worksheet = workbook.create_sheet(
                title=sheet_name[:31]
            )

            headers = sheet_data.get(
                "headers",
                [],
            )

            rows = sheet_data.get(
                "rows",
                [],
            )

            # Headers
            for column_index, header in enumerate(
                    headers,
                    start=1,
            ):
                cell = worksheet.cell(
                    row=1,
                    column=column_index,
                    value=header,
                )

                cell.font = Font(
                    bold=True,
                )

                cell.fill = PatternFill(
                    fill_type="solid",
                    fgColor="D9EAF7",
                )

                cell.alignment = Alignment(
                    horizontal="center",
                )

            # Data
            for row_index, row in enumerate(
                    rows,
                    start=2,
            ):
                for column_index, value in enumerate(
                        row,
                        start=1,
                ):
                    worksheet.cell(
                        row=row_index,
                        column=column_index,
                        value=value,
                    )

            # Auto width
            for column_cells in worksheet.columns:
                max_length = 0

                column_letter = get_column_letter(
                    column_cells[0].column
                )

                for cell in column_cells:
                    value = (
                        str(cell.value)
                        if cell.value is not None
                        else ""
                    )

                    max_length = max(
                        max_length,
                        len(value),
                    )

                worksheet.column_dimensions[
                    column_letter
                ].width = min(
                    max_length + 2,
                    40,
                    )

            worksheet.freeze_panes = "A2"

        workbook.save(output_path)