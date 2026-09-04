from openpyxl import load_workbook


class ExcelParser:

    def extract(self, file_path: str):
        workbook = load_workbook(
            file_path,
            data_only=False,
        )

        sheets = []

        for worksheet in workbook.worksheets:
            rows = []

            for row in worksheet.iter_rows(
                    values_only=True
            ):
                rows.append([
                    value
                    for value in row
                ])

            sheets.append({
                "name": worksheet.title,
                "headers": rows[0] if rows else [],
                "rows": rows[1:] if len(rows) > 1 else [],
            })

        return sheets