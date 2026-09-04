from openpyxl import load_workbook


class ExcelModifier:

    def modify(
            self,
            file_path: str,
            modifications: list[dict],
            output_path: str,
    ):
        workbook = load_workbook(
            file_path
        )

        for modification in modifications:
            action = modification.get("action")
            sheet_name = modification.get("sheet")

            if sheet_name not in workbook.sheetnames:
                continue

            worksheet = workbook[sheet_name]

            if action == "update_cell":
                cell = modification.get("cell")
                value = modification.get("value")

                if cell:
                    worksheet[cell] = value

            elif action == "replace_text":
                old_text = str(
                    modification.get(
                        "old_text",
                        "",
                    )
                )

                new_text = str(
                    modification.get(
                        "new_text",
                        "",
                    )
                )

                for row in worksheet.iter_rows():
                    for cell in row:
                        if (
                                cell.value is not None
                                and old_text in str(cell.value)
                        ):
                            cell.value = str(
                                cell.value
                            ).replace(
                                old_text,
                                new_text,
                            )

            elif action == "add_column":
                column_index = modification.get(
                    "column_index"
                )

                header = modification.get(
                    "header"
                )

                if column_index:
                    worksheet.insert_cols(
                        column_index
                    )

                    worksheet.cell(
                        row=1,
                        column=column_index,
                        value=header,
                    )

                    values = modification.get(
                        "values",
                        [],
                    )

                    for row_index, value in enumerate(
                            values,
                            start=2,
                    ):
                        worksheet.cell(
                            row=row_index,
                            column=column_index,
                            value=value,
                        )

            elif action == "remove_column":
                column_index = modification.get(
                    "column_index"
                )

                if column_index:
                    worksheet.delete_cols(
                        column_index
                    )

            elif action == "add_row":
                values = modification.get(
                    "values",
                    [],
                )

                worksheet.append(values)

            elif action == "remove_row":
                row_index = modification.get(
                    "row_index"
                )

                if row_index and row_index > 1:
                    worksheet.delete_rows(
                        row_index
                    )

            elif action == "create_sheet":
                new_sheet_name = modification.get(
                    "name",
                    "New Sheet",
                )

                if new_sheet_name not in workbook.sheetnames:
                    worksheet = workbook.create_sheet(
                        new_sheet_name
                    )

                    headers = modification.get(
                        "headers",
                        [],
                    )

                    rows = modification.get(
                        "rows",
                        [],
                    )

                    for column_index, header in enumerate(
                            headers,
                            start=1,
                    ):
                        worksheet.cell(
                            row=1,
                            column=column_index,
                            value=header,
                        )

                    for row_index, row in enumerate(
                            rows,
                            start=2,
                    ):
                        for column_index, value in enumerate(
                                row,
                                start=1,
                        ):
                            worksheet.cell(
                                row=row_index,
                                column=column_index,
                                value=value,
                            )

        workbook.save(output_path)